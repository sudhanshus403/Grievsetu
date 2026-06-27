from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from models.user import GrievanceRequest, UserCreate, UserLogin, ProfileUpdate, PasswordUpdate, NotificationCreate, SupportRequest, DeleteAccountRequest
from models.grievance import Grievance
from models.notification import Notification
from models.schema import User
from database.database import engine, SessionLocal, Base
from auth.token import create_jwt, verify_jwt
from services.process import process_grievance
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import os
import shutil
import uuid
import io
import re
from sqlalchemy import func
from datetime import datetime, timedelta
import random
import string

def generate_user_id():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))


app = FastAPI()

# ── CORS Middleware ──────────────────────────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

Base.metadata.create_all(bind=engine)

def get_db():
    
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ── Helper: Create notification ──────────────────────────────────
def create_notification(db, user_id: str, title: str, description: str, ntype: str = "info"):
    notif = Notification(
        user_id=user_id,
        title=title,
        description=description,
        type=ntype,
    )
    db.add(notif)
    db.commit()
    db.refresh(notif)
    return notif

@app.get("/")
def test():
    return {"message": "Hello World"}
    
@app.post("/register")
def register(user: UserCreate, db: SessionLocal = Depends(get_db)):
    existing_user = db.query(User).filter(User.email == user.email).first()
    
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")
        
    new_user = User(
        id=generate_user_id(),
        name = user.name,
        email = user.email,
        password = user.password,
        role = "citizen"
    )
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return {"message": "User registered successfully"}
    
@app.post("/login")
def login(user: UserLogin, db: SessionLocal = Depends(get_db)):
    # --- Purge accounts whose 24-hour grace period has expired ---
    expired_users = db.query(User).filter(
        User.delete_scheduled_at != None,
        User.delete_scheduled_at <= datetime.now()
    ).all()
    for eu in expired_users:
        # Delete all grievances, notifications, uploads for expired user
        db.query(Grievance).filter(Grievance.user_id == eu.id).delete()
        db.query(Notification).filter(Notification.user_id == eu.id).delete()
        db.delete(eu)
    if expired_users:
        db.commit()

    existing_user = db.query(User).filter(User.email == user.email).first()
    
    if not existing_user:
        raise HTTPException(status_code = 404, detail = "User not found")
    
    if existing_user.password != user.password:
        raise HTTPException(status_code = 401, detail = "Invalid password")
    
    # --- If user had a pending deletion, cancel it on login ---
    if existing_user.delete_scheduled_at is not None:
        existing_user.delete_scheduled_at = None
        db.commit()
        create_notification(
            db, existing_user.id,
            "Account Deletion Cancelled",
            "Your account deletion has been cancelled because you logged in. Your data is safe.",
            "success"
        )
    
    return {"message": "User logged in successfully", 
            "token": create_jwt(user.email),
            "user_id" : existing_user.id,
            "role" : existing_user.role,
            "name" : existing_user.name,
            "email": existing_user.email
        }
    
@app.get("/verifylogin")
def verify_login(payload = Depends(verify_jwt)):
    return {"message": "Token is valid", "payload": payload}


@app.post("/predict")
def predict_grievance(request: GrievanceRequest):
    result = process_grievance(
        text = request.text,
        img_path=request.img_path
    )
    return {
        "status": "success",
        "result": result
    }

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename)[1]
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_FOLDER, unique_name)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"file_path": file_path, "filename": file.filename}

# ── Upload profile picture ──────────────────────────────────────
@app.post("/upload_profile_pic/{user_id}")
async def upload_profile_pic(user_id: str, file: UploadFile = File(...), db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    ext = os.path.splitext(file.filename)[1]
    unique_name = f"profile_{user_id}_{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_FOLDER, unique_name)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    user.profile_pic = file_path
    db.commit()
    
    return {"file_path": file_path, "message": "Profile picture updated"}

    
@app.post("/submit")
def submit_grievance(request: GrievanceRequest, db: SessionLocal = Depends(get_db)):
    result = process_grievance(
        text = request.text,
        img_path = request.img_path,
        user_id = request.user_id,
        address = request.address
    )
    print("USER_ID :", request.user_id)
    
    # Create notification for grievance submission
    if "grievance_id" in result:
        create_notification(
            db, request.user_id,
            "Grievance Submitted",
            f'Your grievance (#{result["grievance_id"]}) has been successfully registered. You will be notified when an officer is assigned.',
            "success"
        )
    
    return result

@app.get("/my_grievances/{user_id}")
def my_grievances(user_id: str, db: SessionLocal = Depends(get_db)):
    data = db.query(Grievance).filter(Grievance.user_id == user_id).all()
    return [
        {
            "id": grievance.id,
            "text": grievance.text,
            "category": grievance.category,
            "status": grievance.status,
            "priority": grievance.priority,
            "created_at": grievance.created_at,
            "image_path": grievance.image_path,
            "address": grievance.address or ""
        }
        for grievance in data
    ]

@app.get("/dashboard/{user_id}")
def dashboard(user_id: str, db: SessionLocal = Depends(get_db)):

    total = db.query(Grievance).filter(Grievance.user_id == user_id).count()
    pending = db.query(Grievance).filter(Grievance.user_id == user_id, Grievance.status == "pending").count()
    resolved = db.query(Grievance).filter(Grievance.user_id == user_id, Grievance.status == "resolved").count()
    in_progress = db.query(Grievance).filter(Grievance.user_id == user_id, Grievance.status == "in_progress").count()

    return {
        "total": total,
        "pending": pending,
        "resolved": resolved,
        "in_progress": in_progress
    }

@app.get("/recent/{user_id}")
def recent(user_id: str, db: SessionLocal = Depends(get_db)):
    data = db.query(Grievance).filter(Grievance.user_id == user_id).order_by(Grievance.created_at.desc()).limit(5).all()
    return [
        {
            "id": grievance.id,
            "text": grievance.text,
            "category": grievance.category,
            "status": grievance.status,
        }
        for grievance in data
    ]

@app.get("/report/{user_id}")
def report(user_id: str, db: SessionLocal = Depends(get_db)):
    data = db.query(func.date(Grievance.created_at),func.count()).filter(Grievance.user_id == user_id).group_by(func.date(Grievance.created_at)).all()
    return [
        {
            "date": date,
            "count": count
        }
        for date, count in data
    ]

# ═══════════════════════════════════════════════════════════════
#  PROFILE ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/profile/{user_id}")
def get_profile(user_id: str, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return {
        "id": user.id,
        "name": user.name,
        "email": user.email,
        "phone": user.phone or "",
        "address": user.address or "",
        "profile_pic": user.profile_pic or "",
        "role": user.role
    }

@app.put("/profile/{user_id}")
def update_profile(user_id: str, profile: ProfileUpdate, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    changes = []
    if profile.name and profile.name != user.name:
        user.name = profile.name
        changes.append("Name")
    if profile.phone is not None and profile.phone != (user.phone or ""):
        user.phone = profile.phone
        changes.append("Phone Number")
    if profile.address is not None and profile.address != (user.address or ""):
        user.address = profile.address
        changes.append("Address")
    
    db.commit()
    
    # Create notification for profile changes
    if changes:
        change_list = ", ".join(changes)
        create_notification(
            db, user_id,
            "Profile Updated",
            f"Your {change_list} {'has' if len(changes)==1 else 'have'} been updated successfully.",
            "info"
        )
    
    return {"message": "Profile updated successfully", "changes": changes}

@app.put("/change_password/{user_id}")
def change_password(user_id: str, data: PasswordUpdate, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.password != data.current_password:
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    
    user.password = data.new_password
    db.commit()
    
    # Create notification for password change
    create_notification(
        db, user_id,
        "Password Changed",
        "Your account password was changed successfully. If you didn't make this change, please contact support immediately.",
        "security"
    )
    
    return {"message": "Password updated successfully"}

# ═══════════════════════════════════════════════════════════════
#  NOTIFICATION ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/notifications/{user_id}")
def get_notifications(user_id: str, db: SessionLocal = Depends(get_db)):
    data = db.query(Notification).filter(Notification.user_id == user_id).order_by(Notification.created_at.desc()).all()
    return [
        {
            "id": n.id,
            "title": n.title,
            "description": n.description,
            "type": n.type,
            "is_read": n.is_read,
            "created_at": n.created_at.isoformat() if n.created_at else None
        }
        for n in data
    ]

@app.get("/notifications/unread_count/{user_id}")
def unread_count(user_id: str, db: SessionLocal = Depends(get_db)):
    count = db.query(Notification).filter(Notification.user_id == user_id, Notification.is_read == False).count()
    return {"count": count}

@app.put("/notifications/mark_read/{notif_id}")
def mark_read(notif_id: int, db: SessionLocal = Depends(get_db)):
    notif = db.query(Notification).filter(Notification.id == notif_id).first()
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    notif.is_read = True
    db.commit()
    return {"message": "Notification marked as read"}

@app.put("/notifications/mark_all_read/{user_id}")
def mark_all_read(user_id: str, db: SessionLocal = Depends(get_db)):
    db.query(Notification).filter(Notification.user_id == user_id, Notification.is_read == False).update({"is_read": True})
    db.commit()
    return {"message": "All notifications marked as read"}

@app.delete("/notifications/{notif_id}")
def delete_notification(notif_id: int, db: SessionLocal = Depends(get_db)):
    notif = db.query(Notification).filter(Notification.id == notif_id).first()
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    db.delete(notif)
    db.commit()
    return {"message": "Notification deleted"}

# Admin send notification to user
@app.post("/notifications/send")
def send_notification(data: NotificationCreate, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == data.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    create_notification(db, data.user_id, data.title, data.description, data.type)
    return {"message": "Notification sent successfully"}

# ═══════════════════════════════════════════════════════════════
#  EXPORT GRIEVANCES AS EXCEL
# ═══════════════════════════════════════════════════════════════

@app.get("/export/{user_id}")
def export_grievances(user_id: str, db: SessionLocal = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    
    data = db.query(Grievance).filter(Grievance.user_id == user_id).order_by(Grievance.created_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Grievances"
    
    # Header styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Grievance ID", "Title/Description", "Category", "Status", "Image", "Date & Time"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    for row_idx, g in enumerate(data, 2):
        values = [
            g.id,
            g.text,
            g.category,
            g.status,
            g.image_path or "No Image",
            g.created_at.strftime("%d/%m/%Y %I:%M %p") if g.created_at else "N/A"
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 22
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=GrievSetu_Grievances.xlsx"}
    )

# ═══════════════════════════════════════════════════════════════
#  SUPPORT / HELP ENDPOINT
# ═══════════════════════════════════════════════════════════════

@app.post("/support")
def submit_support(data: SupportRequest):
    # In production, this would send an email or create a support ticket
    return {"message": "Support request submitted successfully. Our team will contact you shortly."}

# ═══════════════════════════════════════════════════════════════
#  LOGOUT ALL DEVICES
# ═══════════════════════════════════════════════════════════════

@app.post("/logout_all_devices/{user_id}")
def logout_all_devices(user_id: str, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Increment token version — all existing JWTs become stale
    user.token_version = (user.token_version or 0) + 1
    db.commit()
    
    create_notification(
        db, user_id,
        "All Sessions Logged Out",
        "All active sessions on other devices have been logged out for security.",
        "security"
    )
    
    return {"message": "All other sessions have been logged out", "new_token_version": user.token_version}

# ═══════════════════════════════════════════════════════════════
#  ACCOUNT DELETION (24-hour grace period)
# ═══════════════════════════════════════════════════════════════

@app.post("/request_delete/{user_id}")
def request_delete_account(user_id: str, data: DeleteAccountRequest, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify password
    if user.password != data.password:
        raise HTTPException(status_code=401, detail="Incorrect password")
    
    if user.delete_scheduled_at is not None:
        raise HTTPException(status_code=400, detail="Account deletion already scheduled")
    
    # Schedule deletion 24 hours from now
    user.delete_scheduled_at = datetime.now() + timedelta(hours=24)
    db.commit()
    
    # Notify the user
    create_notification(
        db, user_id,
        "Account Deletion Scheduled",
        "Your account has been scheduled for permanent deletion in 24 hours. Log in again to cancel this action and save your account.",
        "warning"
    )
    
    return {
        "message": "Account deletion scheduled. You have 24 hours to log back in to cancel.",
        "delete_scheduled_at": user.delete_scheduled_at.isoformat()
    }

@app.post("/cancel_delete/{user_id}")
def cancel_delete_account(user_id: str, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.delete_scheduled_at is None:
        raise HTTPException(status_code=400, detail="No pending deletion to cancel")
    
    user.delete_scheduled_at = None
    db.commit()
    
    create_notification(
        db, user_id,
        "Account Deletion Cancelled",
        "Your account deletion has been cancelled. Your data is safe.",
        "success"
    )
    
    return {"message": "Account deletion cancelled successfully"}

@app.get("/delete_status/{user_id}")
def get_delete_status(user_id: str, db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "delete_scheduled": user.delete_scheduled_at is not None,
        "delete_scheduled_at": user.delete_scheduled_at.isoformat() if user.delete_scheduled_at else None
    }


# ═══════════════════════════════════════════════════════════════
#  ADMIN ROUTES (merged from Admin_Backend)
# ═══════════════════════════════════════════════════════════════

def admin_required(payload = Depends(verify_jwt), db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.email == payload["email"]).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

@app.post("/admin/register")
def admin_register(user: UserCreate, db: SessionLocal = Depends(get_db)):
    existing_user = db.query(User).filter(User.email == user.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")
    new_user = User(
        id=generate_user_id(),
        name=user.name,
        email=user.email,
        password=user.password,
        role="admin"
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"message": "Admin registered successfully"}

@app.post("/admin/login")
def admin_login(user: UserLogin, db: SessionLocal = Depends(get_db)):
    existing_user = db.query(User).filter(User.email == user.email).first()
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    if existing_user.password != user.password:
        raise HTTPException(status_code=401, detail="Invalid password")
    if existing_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not an admin account")
    return {
        "message": "Admin logged in successfully",
        "token": create_jwt(user.email),
        "user_id": existing_user.id,
        "role": existing_user.role,
        "name": existing_user.name
    }

class AssignAdmin(BaseModel):
    email: str

@app.post("/admin/assign")
def assign_admin(data: AssignAdmin, admin=Depends(admin_required), db: SessionLocal = Depends(get_db)):
    user = db.query(User).filter(User.email == data.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.role = "admin"
    db.commit()
    return {"message": f"{data.email} is now an admin"}

@app.get("/admin/grievances")
def get_grievances(
    date_filter: Optional[str] = Query(None),
    custom_start: Optional[str] = Query(None),
    custom_end: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=500),
    db: SessionLocal = Depends(get_db)
):
    query = db.query(Grievance)
    now = datetime.now()
    if date_filter == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        query = query.filter(Grievance.created_at >= start)
    elif date_filter == "yesterday":
        start = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=0, minute=0, second=0, microsecond=0)
        query = query.filter(Grievance.created_at >= start, Grievance.created_at < end)
    elif date_filter == "week":
        start = now - timedelta(days=7)
        query = query.filter(Grievance.created_at >= start)
    elif date_filter == "month":
        start = now - timedelta(days=30)
        query = query.filter(Grievance.created_at >= start)
    elif date_filter == "custom" and custom_start and custom_end:
        try:
            s = datetime.strptime(custom_start, "%Y-%m-%d")
            e = datetime.strptime(custom_end, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Grievance.created_at >= s, Grievance.created_at < e)
        except ValueError:
            pass

    total = query.count()
    data = query.order_by(Grievance.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    results = []
    for g in data:
        user = db.query(User).filter(User.id == g.user_id).first()
        results.append({
            "id": g.id,
            "user_id": g.user_id,
            "user_name": user.name if user else "Unknown",
            "user_phone": getattr(user, 'phone', '') if user else "",
            "text": g.text,
            "category": g.category,
            "status": g.status,
            "priority": g.priority,
            "department": g.department,
            "image_path": g.image_path,
            "created_at": g.created_at.isoformat() if g.created_at else None,
            "confidence_score": g.confidence_score,
            "address": g.address or ""
        })

    return {
        "grievances": results,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page
    }

class StatusUpdate(BaseModel):
    status: str

@app.put("/admin/update_status/{id}")
def update_status(id: int, data: StatusUpdate, db: SessionLocal = Depends(get_db)):
    grievance = db.query(Grievance).filter(Grievance.id == id).first()
    if not grievance:
        raise HTTPException(status_code=404, detail="Grievance not found")

    grievance.status = data.status
    db.commit()

    status_labels = {
        'pending': 'Pending', 'in_progress': 'In Progress',
        'resolved': 'Resolved', 'escalated': 'Escalated'
    }
    new_label = status_labels.get(data.status, data.status)

    notif_map = {
        'in_progress': ("Grievance In Progress", f"Your grievance (#{grievance.id}) is now being reviewed.", "info"),
        'resolved': ("Grievance Resolved", f"Your grievance (#{grievance.id}) has been resolved.", "success"),
        'escalated': ("Grievance Escalated", f"Your grievance (#{grievance.id}) has been escalated.", "warning"),
    }
    if data.status in notif_map:
        title, desc, ntype = notif_map[data.status]
        create_notification(db, grievance.user_id, title, desc, ntype)

    return {"message": f"Status updated to {new_label} and user notified"}

class PriorityUpdate(BaseModel):
    priority: str

@app.put("/admin/update_priority/{id}")
def update_priority(id: int, data: PriorityUpdate, db: SessionLocal = Depends(get_db)):
    grievance = db.query(Grievance).filter(Grievance.id == id).first()
    if not grievance:
        raise HTTPException(status_code=404, detail="Grievance not found")
    grievance.priority = data.priority
    db.commit()
    return {"message": "Priority updated successfully"}

class EscalateRequest(BaseModel):
    reason: str = ""

@app.put("/admin/escalate/{id}")
def escalate_grievance(id: int, data: EscalateRequest, db: SessionLocal = Depends(get_db)):
    grievance = db.query(Grievance).filter(Grievance.id == id).first()
    if not grievance:
        raise HTTPException(status_code=404, detail="Grievance not found")
    grievance.priority = "critical"
    grievance.status = "escalated"
    db.commit()
    create_notification(db, grievance.user_id, "Grievance Escalated",
        f"Your grievance (#{grievance.id}) has been escalated to higher authorities.", "warning")
    return {"message": "Grievance escalated to higher authority"}

@app.get("/admin/export_grievances")
def admin_export_grievances(
    date_filter: Optional[str] = Query(None),
    custom_start: Optional[str] = Query(None),
    custom_end: Optional[str] = Query(None),
    db: SessionLocal = Depends(get_db)
):
    import pandas as pd
    query = db.query(
        Grievance.id, Grievance.user_id, User.name.label("user_name"),
        User.email.label("user_email"), User.phone.label("user_phone"),
        Grievance.category, Grievance.text, Grievance.status, Grievance.created_at
    ).join(User, Grievance.user_id == User.id)

    now = datetime.now()
    if date_filter == 'today':
        query = query.filter(Grievance.created_at >= now.replace(hour=0, minute=0, second=0, microsecond=0))
    elif date_filter == 'yesterday':
        yesterday = now - timedelta(days=1)
        query = query.filter(Grievance.created_at >= yesterday.replace(hour=0, minute=0, second=0, microsecond=0),
                             Grievance.created_at < now.replace(hour=0, minute=0, second=0, microsecond=0))
    elif date_filter == 'week':
        query = query.filter(Grievance.created_at >= now - timedelta(days=7))
    elif date_filter == 'month':
        query = query.filter(Grievance.created_at >= now - timedelta(days=30))
    elif date_filter == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d')
        end_date = datetime.strptime(custom_end, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Grievance.created_at >= start_date, Grievance.created_at < end_date)

    grievances = query.all()
    data = []
    for g in grievances:
        text = g.text or ""
        cleaned = re.sub(r'^\[.*?\]\s*', '', text)
        dot_idx = cleaned.find('.')
        if dot_idx != -1:
            title = cleaned[:dot_idx].strip()
            desc = cleaned[dot_idx+1:].strip() or 'No description provided'
        else:
            title = cleaned.strip()
            desc = 'No description provided'
        data.append({
            "Grievance ID": g.id,
            "User Name with ID": f"{g.user_name} (UID: {g.user_id})",
            "Title": title, "Description": desc,
            "Category": g.category, "Status": g.status,
            "Date Submitted": g.created_at.strftime("%Y-%m-%d %H:%M:%S") if g.created_at else ""
        })

    df = pd.DataFrame(data)
    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Grievances')
    stream.seek(0)
    return StreamingResponse(stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=grievances_export_{now.strftime('%Y%m%d')}.xlsx"})

@app.get("/admin/stats")
def admin_stats(db: SessionLocal = Depends(get_db)):
    total = db.query(Grievance).count()
    pending = db.query(Grievance).filter(Grievance.status == "pending").count()
    in_progress = db.query(Grievance).filter(Grievance.status == "in_progress").count()
    resolved = db.query(Grievance).filter(Grievance.status == "resolved").count()
    escalated = db.query(Grievance).filter(Grievance.status == "escalated").count()
    today_count = db.query(Grievance).filter(
        Grievance.created_at >= datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    return {"total": total, "pending": pending, "in_progress": in_progress,
            "resolved": resolved, "escalated": escalated, "today": today_count}

class NotifSend(BaseModel):
    user_id: str
    title: str
    description: str
    type: str = "admin"

@app.post("/admin/send_notification")
def admin_send_notification(data: NotifSend, db: SessionLocal = Depends(get_db)):
    create_notification(db, data.user_id, data.title, data.description, data.type)
    return {"message": "Notification sent"}

@app.get("/admin/pending_deletions")
def pending_deletions(db: SessionLocal = Depends(get_db)):
    users = db.query(User).filter(User.delete_scheduled_at != None).all()
    return [
        {
            "id": u.id, "name": u.name, "email": u.email,
            "delete_scheduled_at": u.delete_scheduled_at.isoformat() if u.delete_scheduled_at else None,
            "hours_remaining": max(0, round((u.delete_scheduled_at - datetime.now()).total_seconds() / 3600, 1)) if u.delete_scheduled_at else 0
        }
        for u in users
    ]


# Serve uploaded files as static (must be after all API routes)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
